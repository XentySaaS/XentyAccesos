"""ViewSet de Empleado (contexto proveedores) + import por Excel idempotente."""

from __future__ import annotations

from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.response import Response

from apps.config.services import AuditViewSetMixin
from common.permissions import PERMISOS_BASE, ContextoProveedores, RequiereModulo
from common.validators import validar_archivo

from .models import Empleado
from .serializers import EmpleadoSerializer


class EmpleadoViewSet(AuditViewSetMixin, viewsets.ModelViewSet):
    serializer_class = EmpleadoSerializer
    permission_classes = [*PERMISOS_BASE(), ContextoProveedores, RequiereModulo("empleados")]
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    filterset_fields = ["estado"]

    def get_queryset(self):
        actor = self.request.user
        qs = Empleado.objects.all().order_by("id")
        # El proveedor solo ve los empleados de SU empresa (todas las cuentas del mismo Proveedor).
        if getattr(actor, "proveedor_id", None):
            return qs.filter(proveedor__proveedor_id=actor.proveedor_id)
        return qs.filter(proveedor=actor)

    def perform_create(self, serializer):
        serializer.validated_data["proveedor"] = self.request.user
        super().perform_create(serializer)

    @action(detail=False, methods=["get"])
    def plantilla(self, request):
        """Descarga la plantilla .xlsx para importar (mismos encabezados que lee ``importar``)."""
        from io import BytesIO

        from django.http import HttpResponse
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Empleados"
        # Fila 1 = encabezados: DEBE coincidir con el orden que consume ``importar`` (nombre, email, telefono).
        ws.append(["nombre", "email", "telefono"])
        # Fila de ejemplo para que el usuario entienda el formato esperado (email obligatorio; 10 dígitos).
        ws.append(["Juan Pérez López", "juan.perez@ejemplo.com", "5512345678"])
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 16

        buffer = BytesIO()
        wb.save(buffer)
        resp = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="plantilla-empleados.xlsx"'
        return resp

    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser, FormParser])
    def importar(self, request):
        """Importa empleados desde un .xlsx (columnas: nombre, email, telefono). Idempotente por email."""
        archivo = request.FILES.get("archivo")
        if not archivo:
            return Response(
                {"detail": "Falta 'archivo' (.xlsx)."}, status=status.HTTP_400_BAD_REQUEST
            )
        validar_archivo(archivo, extensiones=(".xlsx",), max_mb=5)

        from openpyxl import load_workbook

        from common.phone import normalizar_telefono

        empresa_id = request.user.proveedor_id
        ws = load_workbook(archivo, read_only=True, data_only=True).active
        creados = actualizados = omitidos = 0
        for fila in ws.iter_rows(min_row=2, values_only=True):
            if not fila or not fila[0]:
                continue
            nombre = str(fila[0]).strip()
            email = str(fila[1]).strip().lower() if len(fila) > 1 and fila[1] else ""
            telefono = normalizar_telefono(str(fila[2])) if len(fila) > 2 and fila[2] else ""
            # El correo es la llave de deduplicación (y ahora obligatorio): sin él, se omite la fila.
            if not email:
                omitidos += 1
                continue
            # Deduplica a nivel EMPRESA (los empleados se comparten entre cuentas del mismo Proveedor).
            existente = (
                Empleado.objects.filter(proveedor__proveedor_id=empresa_id, email__iexact=email)
                .exclude(estado=Empleado.Estado.BAJA)
                .first()
            )
            if existente:
                existente.nombre = nombre
                if telefono:
                    existente.telefono = telefono
                existente.save(update_fields=["nombre", "telefono", "actualizado"])
                actualizados += 1
            else:
                Empleado.objects.create(
                    proveedor=request.user, email=email, nombre=nombre, telefono=telefono or None
                )
                creados += 1
        return Response({"creados": creados, "actualizados": actualizados, "omitidos": omitidos})
