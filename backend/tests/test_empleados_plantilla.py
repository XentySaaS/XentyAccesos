"""Plantilla de importación de empleados: es un .xlsx válido cuyos encabezados coinciden
con lo que consume ``EmpleadoViewSet.importar`` (nombre, email, telefono).

La acción ``plantilla`` no toca BD ni usa el usuario: se puede invocar directamente sobre el
ViewSet con una request cualquiera.
"""

from __future__ import annotations

from io import BytesIO

from rest_framework.request import Request
from rest_framework.test import APIRequestFactory

_factory = APIRequestFactory()


def _plantilla_bytes() -> bytes:
    from apps.empleados.views import EmpleadoViewSet

    resp = EmpleadoViewSet().plantilla(Request(_factory.get("/api/empleados/plantilla/")))
    assert resp.status_code == 200
    assert "spreadsheetml" in resp["Content-Type"]
    assert "attachment" in resp["Content-Disposition"]
    assert ".xlsx" in resp["Content-Disposition"]
    return resp.content


def test_plantilla_es_xlsx_valido_con_encabezados_correctos():
    from openpyxl import load_workbook

    ws = load_workbook(BytesIO(_plantilla_bytes())).active
    # Fila 1 = encabezados en el MISMO orden que lee ``importar``.
    assert [c.value for c in ws[1]] == ["nombre", "email", "telefono"]
    # Fila 2 = ejemplo con email presente (el email es obligatorio en el importador).
    ejemplo = [c.value for c in ws[2]]
    assert ejemplo[0] and "@" in str(ejemplo[1])
